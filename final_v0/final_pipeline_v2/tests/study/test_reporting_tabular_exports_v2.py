from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from ppg_frailty.reporting.tabular import (
    ReportTable,
    compact_rows,
    write_csv,
    write_excel_workbook,
    write_excel_workbook_from_csv_directory,
)


class ReportingTabularExportTests(unittest.TestCase):
    def test_compact_rows_collapses_score_mean_and_sd_without_losing_json_source(self) -> None:
        source = [
            {
                "case_id": "case-a",
                "metric": "balanced_accuracy",
                "n": 5,
                "mean": 0.726,
                "sample_sd": 0.060,
                "population_sd": 0.054,
                "ci95_low": 0.65,
                "ci95_high": 0.80,
            }
        ]

        displayed = compact_rows(source)

        self.assertEqual(
            displayed,
            [
                {
                    "case_id": "case-a",
                    "metric": "balanced_accuracy",
                    "n": 5,
                    "mean_sd": "72.6 ± 6.0",
                    "ci95": "[65.0, 80.0]",
                }
            ],
        )
        self.assertEqual(source[0]["mean"], 0.726)
        self.assertIn("ci95_low", source[0])

    def test_workbook_has_one_valid_xml_worksheet_per_registered_table(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = write_excel_workbook(
                Path(temporary) / "report_tables.xlsx",
                (
                    ReportTable(
                        "metric_distribution_summary",
                        [{"metric": "macro_f1", "mean": 0.8, "sample_sd": 0.1}],
                    ),
                    ReportTable("empty_evidence_table", ()),
                ),
            )

            with ZipFile(path) as archive:
                self.assertIsNone(archive.testzip())
                worksheets = [
                    name
                    for name in archive.namelist()
                    if name.startswith("xl/worksheets/sheet")
                ]
                self.assertEqual(len(worksheets), 2)
                for name in archive.namelist():
                    if name.endswith((".xml", ".rels")):
                        ElementTree.fromstring(archive.read(name))

    def test_workbook_is_readable_by_installed_openpyxl_when_available(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as temporary:
            target = Path(temporary) / "tables.xlsx"
            write_excel_workbook(
                target,
                (ReportTable("metrics", ({"case": "a", "score": 0.75},)),),
            )
            workbook = load_workbook(target, read_only=True, data_only=True)
            self.assertEqual(workbook.sheetnames, ["metrics"])
            self.assertEqual(workbook["metrics"]["A2"].value, "a")

    def test_csv_directory_workbook_includes_late_table_figure_pairs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            tables = Path(temporary) / "tables"
            tables.mkdir()
            write_csv(tables / "case_summary.csv", ({"case_id": "a"},))
            write_csv(tables / "repeat_metrics.csv", ({"repeat": 0},))
            # This registry is deliberately written last, matching the generic
            # report generation order that previously omitted it from XLSX.
            write_csv(
                tables / "table_figure_pairs.csv",
                ({"table": "case_summary", "figure": "leaderboard"},),
            )
            nested = tables / "top_confusion_matrices"
            nested.mkdir()
            write_csv(nested / "auxiliary.csv", ({"value": 1},))

            workbook_path = write_excel_workbook_from_csv_directory(
                tables / "report_tables.xlsx",
                tables,
            )

            with ZipFile(workbook_path) as archive:
                workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            namespace = {
                "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            }
            sheets = workbook.findall(".//m:sheet", namespace)
            self.assertEqual(len(sheets), len(tuple(tables.glob("*.csv"))))
            self.assertEqual(
                {sheet.attrib["name"] for sheet in sheets},
                {"case_summary", "repeat_metrics", "table_figure_pairs"},
            )

    def test_csv_directory_workbook_preserves_field_larger_than_128_kib(self) -> None:
        large_value = '{"payload":"' + ("x" * 140_000) + '"}'
        previous_limit = csv.field_size_limit()
        try:
            csv.field_size_limit(128 * 1024)
            with tempfile.TemporaryDirectory() as temporary:
                tables = Path(temporary) / "tables"
                tables.mkdir()
                write_csv(
                    tables / "test_components.csv",
                    (
                        {
                            "component_id": "large-json-fixture",
                            "fixed_parameters": large_value,
                        },
                    ),
                )

                workbook_path = write_excel_workbook_from_csv_directory(
                    tables / "report_tables.xlsx",
                    tables,
                )

                self.assertEqual(csv.field_size_limit(), 128 * 1024)
                with ZipFile(workbook_path) as archive:
                    worksheet = ElementTree.fromstring(
                        archive.read("xl/worksheets/sheet1.xml")
                    )
                namespace = {
                    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                }
                rows = worksheet.findall(".//m:row", namespace)

                def values(row: ElementTree.Element) -> list[str]:
                    return [
                        "".join(
                            text.text or ""
                            for text in cell.findall(".//m:t", namespace)
                        )
                        for cell in row.findall("m:c", namespace)
                    ]

                headers = values(rows[0])
                persisted = values(rows[1])
                field_columns = [
                    index
                    for index, name in enumerate(headers)
                    if name == "fixed_parameters"
                    or name.startswith("fixed_parameters__continuation_")
                ]
                self.assertGreater(len(field_columns), 4)
                self.assertTrue(
                    all(len(persisted[index]) <= 32_767 for index in field_columns)
                )
                self.assertEqual(
                    "".join(persisted[index] for index in field_columns),
                    large_value,
                )
        finally:
            csv.field_size_limit(previous_limit)


if __name__ == "__main__":
    unittest.main()
